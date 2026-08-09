"""Baseline deterministic extractor for spreadsheet statement exports.

openpyxl parses the container and the sheet. What is domain logic, and therefore
what this module actually contains, is three things: finding the field row under
a merged title block and a group row, mapping header synonyms onto field names,
and turning a numeric cell into exact integer cents.

The tabular baseline is imported rather than restated. Header matching, value
parsing, the totals rule, and the resource ceilings are the CSV extractor's, so a
spreadsheet and a delimited export of the same statement are read by one set of
rules and cannot drift apart. Only the reader and the route differ.
"""

import logging
from collections import deque
from collections.abc import Iterator, Sequence
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from crossfoot.constants import (
    FIELD_FAMILIES,
    ExtractionRoute,
    FieldFamily,
    FieldName,
    FieldSource,
    IngestErrorKind,
    ReviewStatus,
)
from crossfoot.extraction.normalize import (
    CENTS_PER_DOLLAR,
    parse_amount_to_cents,
    strip_control_chars,
)
from crossfoot.extraction.tabular import (
    _HEADER_MATCH_RATIO,
    _SYNONYM_TO_FIELD,
    MAX_CELL_CHARS,
    MAX_DATA_ROWS,
    MAX_FILE_BYTES,
    MAX_HEADER_SCAN_ROWS,
    _file_size,
    _grammar_signal,
    _header_key,
    _is_non_data,
    _map_columns,
    _parse_value,
    _unprocessable,
)
from crossfoot.models.extraction import ExtractedDocument, ExtractedField, FieldSignals

_LOGGER = logging.getLogger(__name__)

# Header rows a spreadsheet stacks over its data: a merged group row above the
# field row. Deeper stacks are not a shape this generator or these exports use.
HEADER_ROW_DEPTH = 2
# Columns a header row must occupy. A merged title block collapses to a single
# anchored cell, so a one-column row is decoration however well its words read.
MIN_HEADER_COLUMNS = 2

# Joins the stacked fragments of one column into a single candidate spelling.
_HEADER_FRAGMENT_JOIN = " "

# Currency number formats show a grouped amount behind a symbol, and raw_text is
# contractually what the source displays. Only this one format shape is
# reproduced; the rest of the Excel format language is the library's business.
_CURRENCY_FORMAT_MARKER = "$"


class _CellTooLargeError(Exception):
    """One cell over the shared cell ceiling, raised so the document is refused."""


def extract_xlsx(path: Path, doc_id: str) -> ExtractedDocument:
    """Extract a workbook statement export into typed line fields; never raises."""
    size = _file_size(path)
    if size is None:
        return _unprocessable(path, doc_id, IngestErrorKind.UNRECOGNIZED, "unreadable file")
    if size > MAX_FILE_BYTES:
        # Checked by stat, so a decompression bomb is refused before openpyxl
        # opens the container and inflates whatever is inside it.
        return _unprocessable(
            path,
            doc_id,
            IngestErrorKind.TOO_LARGE,
            f"file is {size} bytes, over the {MAX_FILE_BYTES} byte limit",
        )
    try:
        # read_only streams the sheet instead of materializing it; data_only
        # takes cached values so no formula is ever evaluated here.
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:  # the backing XML parser decides the type
        _LOGGER.warning("cannot open %s as a workbook: %s", path, error)
        return _unprocessable(
            path, doc_id, IngestErrorKind.UNRECOGNIZED, f"unreadable workbook: {error}"
        )
    try:
        column_map, line_fields, over_row_cap = _read_sheet(book, doc_id)
    except _CellTooLargeError as error:
        return _unprocessable(path, doc_id, IngestErrorKind.TOO_LARGE, str(error))
    except Exception as error:  # one workbook must not end a run, whatever it holds
        _LOGGER.warning("%s failed to read: %s", doc_id, error)
        return _unprocessable(
            path, doc_id, IngestErrorKind.UNRECOGNIZED, f"malformed workbook: {error}"
        )
    finally:
        _close(book)
    if column_map is None:
        return _unprocessable(
            path, doc_id, IngestErrorKind.UNRECOGNIZED, "no recognizable header row"
        )
    if over_row_cap:
        return _unprocessable(
            path, doc_id, IngestErrorKind.TOO_LARGE, f"more than {MAX_DATA_ROWS} data rows"
        )
    # Statement totals sit in a stray TOTALS row that is not data, so no
    # crossfoot check here, exactly as in the CSV path.
    return ExtractedDocument(
        doc_id=doc_id,
        file_path=path.as_posix(),
        route=ExtractionRoute.XLSX,
        line_fields=line_fields,
    )


def _read_sheet(
    book: Any, doc_id: str
) -> tuple[dict[int, FieldName] | None, tuple[ExtractedField, ...], bool]:
    """Header map, line fields, and whether the row cap was passed."""
    rows = _sheet_rows(book)
    column_map = _find_header(rows)
    if column_map is None:
        return None, (), False
    line_fields, over_row_cap = _extract_line_fields(rows, column_map, doc_id)
    return column_map, line_fields, over_row_cap


def _close(book: Any) -> None:
    """read_only holds the container open, and Windows blocks callers until it closes."""
    try:
        book.close()
    except Exception as error:  # a failed close must not mask a good extraction
        _LOGGER.warning("cannot close workbook: %s", error)


def _sheet_rows(book: Any) -> Iterator[list[str]]:
    """Every row of the first worksheet as cell text, in sheet order.

    First sheet only: a statement export writes one, and a workbook hiding its
    statement on sheet four is a different problem than this extractor solves.
    """
    sheets = book.worksheets
    if not sheets:
        return
    for row in sheets[0].iter_rows():
        yield [_cell_text(cell.value, getattr(cell, "number_format", "")) for cell in row]


def _cell_text(value: object, number_format: str) -> str:
    """One cell as text, taking every numeric value through Decimal, never a float.

    openpyxl hands back a float for a currency cell. str() of that float is its
    shortest round-tripping literal, so Decimal(str(value)) recovers the digits
    the sheet stores and 1234.55 stays 1234.55 rather than drifting into binary.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        text = value
    elif isinstance(value, bool):
        text = str(value)
    elif isinstance(value, datetime):
        text = value.date().isoformat()
    elif isinstance(value, date):
        text = value.isoformat()
    elif isinstance(value, int | float | Decimal):
        text = _numeric_text(value, number_format)
    else:
        text = str(value)
    if len(text) > MAX_CELL_CHARS:
        raise _CellTooLargeError(f"a cell holds {len(text)} characters, over {MAX_CELL_CHARS}")
    # Stripped here so nothing downstream, header matching included, sees a NUL.
    return strip_control_chars(text)


def _numeric_text(value: int | float | Decimal, number_format: str) -> str:
    """A numeric cell as the sheet displays it, formatted out of exact cents."""
    plain = _plain_decimal_text(value)
    if _CURRENCY_FORMAT_MARKER not in number_format:
        return plain
    cents = parse_amount_to_cents(plain)
    if cents is None:
        return plain
    sign = "-" if cents < 0 else ""
    dollars, remainder = divmod(abs(cents), CENTS_PER_DOLLAR)
    return f"{sign}{_CURRENCY_FORMAT_MARKER}{dollars:,}.{remainder:02d}"


def _plain_decimal_text(value: int | float | Decimal) -> str:
    """Positional decimal digits, never scientific notation the parsers reject."""
    return format(Decimal(str(value)), "f")


def _find_header(rows: Iterator[list[str]]) -> dict[int, FieldName] | None:
    """Consume rows to the first synonym-bearing one, flattening the stack above it.

    Each column of the candidate row is joined with the fragments standing over
    it, so a field name split across the stack still matches. Every fragment is
    also tried alone, bottom up, because a group label such as "References" must
    not hide the field row beneath it and the field row is the more specific one.
    """
    stack: deque[list[str]] = deque(maxlen=HEADER_ROW_DEPTH)
    for _ in range(MAX_HEADER_SCAN_ROWS):
        row = next(rows, None)
        if row is None:
            return None
        stack.append(row)
        column_map = _match_header(_flatten(stack))
        if column_map is not None:
            return column_map
    return None


def _flatten(stack: Sequence[list[str]]) -> list[list[str]]:
    """The stacked header rows as one non-empty fragment list per column."""
    width = max((len(row) for row in stack), default=0)
    return [
        [row[column].strip() for row in stack if column < len(row) and row[column].strip()]
        for column in range(width)
    ]


def _match_header(columns: Sequence[list[str]]) -> dict[int, FieldName] | None:
    """Map the flattened columns when enough of them name a field, else None."""
    pairs = [
        (_HEADER_FRAGMENT_JOIN.join(fragments), _match_key(fragments)) for fragments in columns
    ]
    occupied = [key for text, key in pairs if text]
    if len(occupied) < MIN_HEADER_COLUMNS:
        return None
    matched = sum(1 for key in occupied if key in _SYNONYM_TO_FIELD)
    if matched / len(occupied) < _HEADER_MATCH_RATIO:
        return None
    return _map_columns(pairs)


def _match_key(fragments: Sequence[str]) -> str:
    """First spelling of a stacked header cell that names a field, else the joined text."""
    joined = _HEADER_FRAGMENT_JOIN.join(fragments)
    for candidate in (joined, *reversed(fragments)):
        key = _header_key(candidate)
        if key in _SYNONYM_TO_FIELD:
            return key
    return _header_key(joined)


def _extract_line_fields(
    data_rows: Iterator[list[str]], column_map: dict[int, FieldName], doc_id: str
) -> tuple[tuple[ExtractedField, ...], bool]:
    """Fields for every data row, plus a flag for the row cap being exceeded."""
    fields: list[ExtractedField] = []
    line_no = 0
    for row in data_rows:
        # Blank rows and the stray TOTALS row are not data and take no ordinal,
        # so line_no stays aligned with the statement's own line numbering.
        if _is_non_data(row):
            continue
        if line_no >= MAX_DATA_ROWS:
            return tuple(fields), True
        line_no += 1
        for column, name in column_map.items():
            if column < len(row):
                fields.append(_build_field(doc_id, line_no, name, row[column]))
    return tuple(fields), False


def _build_field(doc_id: str, line_no: int, name: FieldName, text: str) -> ExtractedField:
    """The CSV baseline's field, tagged with the spreadsheet tier.

    Signals and confidence are deliberately the CSV rule: a validator pass is the
    only evidence a deterministic reader has, and the calibrated scorer replaces
    this flat confidence downstream. The reader already stripped control chars.
    """
    family = FIELD_FAMILIES[name]
    try:
        value, value_cents, value_date = _parse_value(family, text)
        grammar = _grammar_signal(name, text.strip()) if family is FieldFamily.REFERENCE else None
    except Exception as error:  # one hostile cell must not void the document
        _LOGGER.warning("%s line %d %s failed to parse: %s", doc_id, line_no, name, error)
        value, value_cents, value_date, grammar = None, None, None, None
    parsed = value is not None
    return ExtractedField(
        field_id=f"fld-{doc_id}-{line_no:04d}-{name}",
        doc_id=doc_id,
        line_no=line_no,
        name=name,
        family=family,
        raw_text=text,
        value=value,
        value_cents=value_cents,
        value_date=value_date,
        source=FieldSource.DETERMINISTIC,
        signals=FieldSignals(
            validator_pass=1.0 if parsed else 0.0,
            grammar_match=grammar,
            route=ExtractionRoute.XLSX,
        ),
        confidence=1.0 if parsed else 0.0,
        status=ReviewStatus.AUTO_ACCEPTED if parsed else ReviewStatus.NEEDS_REVIEW,
    )
