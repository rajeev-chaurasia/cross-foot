"""CSV and XLSX statement renderers with seeded, deterministic messiness."""

import csv
import random
import re
import zipfile
from datetime import date, datetime
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from crossfoot.constants import CSV_HEADER_SYNONYMS, FieldName
from crossfoot.generator.renderers.base import (
    CENTS_PER_DOLLAR,
    DOC_LINE_FIELDS,
    DOC_TITLES,
    LINE_REFERENCE_FIELDS,
    MARQUE_BRANDING,
    grouped_amount,
    header_key,
    line_key,
    line_reference,
)
from crossfoot.models.statement import StatementDoc, StatementLine


class AmountStyle(StrEnum):
    PLAIN = "plain"
    DOLLAR_COMMA = "dollar_comma"
    PAREN_NEGATIVE = "paren_negative"


class DateStyle(StrEnum):
    US_SLASH = "us_slash"  # MM/DD/YYYY
    ISO = "iso"  # YYYY-MM-DD


CSV_DELIMITERS: tuple[str, ...] = (",", "|", "\t")
CSV_ENCODINGS: tuple[str, ...] = ("utf-8", "cp1252")

XLSX_SHEET_TITLE = "Statement"
XLSX_CURRENCY_FORMAT = '"$"#,##0.00'
# Ratio of data rows whose amount lands in a currency-formatted numeric cell.
XLSX_CURRENCY_CELL_RATIO = 0.5
XLSX_TITLE_FONT_SIZE = 14
# Fixed workbook metadata keeps repeated runs comparable; real timestamps never
# belong in dataset artifacts. openpyxl stamps the save time over the properties
# it was handed, and the zip container carries member mtimes of its own, so the
# saved file is rewritten with all three pinned to this instant.
XLSX_FIXED_TIMESTAMP = datetime(2026, 1, 1)
XLSX_FIXED_ZIP_TIME = (2026, 1, 1, 0, 0, 0)
XLSX_CORE_PROPERTIES = "docProps/core.xml"
_CORE_TIMESTAMP_TEXT = b"2026-01-01T00:00:00Z"
_CORE_TIMESTAMP_ELEMENT = re.compile(rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(<)")
XLSX_GROUP_LABELS = ("References", "Activity", "Amounts")
XLSX_TOTALS_LABEL = "TOTALS"
XLSX_COLUMN_WIDTHS: dict[FieldName, float] = {
    FieldName.LINE_DATE: 13.0,
    FieldName.DESCRIPTION: 42.0,
    FieldName.LINE_AMOUNT: 14.0,
    FieldName.CLAIM_NUMBER: 16.0,
    FieldName.RO_NUMBER: 12.0,
    FieldName.VIN: 21.0,
    FieldName.INVOICE_NUMBER: 16.0,
    FieldName.PROGRAM_CODE: 15.0,
}


def format_amount(style: AmountStyle, amount_cents: int) -> str:
    """Amount string in one of the tabular styles; parseable by the extractor."""
    negative = amount_cents < 0
    if style is AmountStyle.PLAIN:
        dollars, cents = divmod(abs(amount_cents), CENTS_PER_DOLLAR)
        magnitude = f"{dollars}.{cents:02d}"
        return f"-{magnitude}" if negative else magnitude
    magnitude = grouped_amount(amount_cents)
    if style is AmountStyle.DOLLAR_COMMA:
        return f"-${magnitude}" if negative else f"${magnitude}"
    return f"({magnitude})" if negative else magnitude


def format_date(style: DateStyle, value: date) -> str:
    if style is DateStyle.ISO:
        return value.isoformat()
    return f"{value.month:02d}/{value.day:02d}/{value.year}"


def _line_cell(
    line: StatementLine, field: FieldName, amount_style: AmountStyle, date_style: DateStyle
) -> str:
    if field is FieldName.LINE_DATE:
        return format_date(date_style, line.line_date)
    if field is FieldName.LINE_AMOUNT:
        return format_amount(amount_style, line.amount_cents)
    if field is FieldName.DESCRIPTION:
        return line.description
    return line_reference(line, field) or ""


def render_csv(doc: StatementDoc, template_id: str, seed: int, out_path: Path) -> dict[str, str]:
    """Line-item CSV with seeded header synonyms, delimiter, encoding, and formats.

    template_id names the variant in the manifest; the CSV layout itself is
    driven entirely by the seed. Header synonyms are consistent within a file.
    """
    rng = random.Random(seed)
    delimiter = rng.choice(CSV_DELIMITERS)
    encoding = rng.choice(CSV_ENCODINGS)
    amount_style = rng.choice(tuple(AmountStyle))
    date_style = rng.choice(tuple(DateStyle))
    fields = DOC_LINE_FIELDS[doc.doc_type]
    header_row = [rng.choice(CSV_HEADER_SYNONYMS[field]) for field in fields]

    rendered: dict[str, str] = {}
    rows: list[list[str]] = [header_row]
    # Contract clarification: data rows in truth line_no order so extractor
    # ordinals align with truth line_no.
    for line in sorted(doc.lines, key=lambda item: item.line_no):
        row: list[str] = []
        for field in fields:
            value = _line_cell(line, field, amount_style, date_style)
            row.append(value)
            if value:
                rendered[line_key(line.line_no, field)] = value
        rows.append(row)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding=encoding, newline="") as handle:
        csv.writer(handle, delimiter=delimiter).writerows(rows)
    return rendered


def _xlsx_field_order(doc: StatementDoc) -> tuple[FieldName, ...]:
    """XLSX groups reference columns first, then activity, then the amount."""
    references = tuple(
        field for field in DOC_LINE_FIELDS[doc.doc_type] if field in LINE_REFERENCE_FIELDS
    )
    return (*references, FieldName.LINE_DATE, FieldName.DESCRIPTION, FieldName.LINE_AMOUNT)


def _write_totals_row(sheet: Any, row_index: int, amount_column: int, subtotal_cents: int) -> int:
    """The stray TOTALS row that real exports love to drop mid-sheet."""
    row_index += 1
    sheet.cell(row=row_index, column=1, value=XLSX_TOTALS_LABEL).font = Font(bold=True)
    totals_cell = sheet.cell(
        row=row_index, column=amount_column, value=subtotal_cents / CENTS_PER_DOLLAR
    )
    totals_cell.number_format = XLSX_CURRENCY_FORMAT
    totals_cell.font = Font(bold=True)
    return row_index


def render_xlsx(doc: StatementDoc, template_id: str, seed: int, out_path: Path) -> dict[str, str]:
    """Workbook with a merged title block, two-row header, and a stray TOTALS row.

    template_id names the variant in the manifest; the sheet layout itself is
    driven entirely by the seed.
    """
    rng = random.Random(seed)
    amount_style = rng.choice(tuple(AmountStyle))
    date_style = rng.choice(tuple(DateStyle))
    fields = _xlsx_field_order(doc)
    synonyms = {field: rng.choice(CSV_HEADER_SYNONYMS[field]) for field in fields}
    totals_at_end = rng.random() < 0.5

    branding = MARQUE_BRANDING[doc.oem]
    statement_date = format_date(date_style, doc.statement_date)
    subtotal_display = format_amount(AmountStyle.DOLLAR_COMMA, doc.subtotal_cents)
    rendered: dict[str, str] = {
        header_key(FieldName.STATEMENT_NUMBER): doc.statement_number,
        header_key(FieldName.STATEMENT_DATE): statement_date,
        header_key(FieldName.SUBTOTAL): subtotal_display,
    }

    book = Workbook()
    book.properties.created = XLSX_FIXED_TIMESTAMP
    book.properties.modified = XLSX_FIXED_TIMESTAMP
    sheet = book.active
    sheet.title = XLSX_SHEET_TITLE
    column_count = len(fields)

    # Merged title block: marque name, then document title.
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = sheet.cell(row=1, column=1, value=branding.name)
    title_cell.font = Font(bold=True, size=XLSX_TITLE_FONT_SIZE)
    title_cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    subtitle_cell = sheet.cell(row=2, column=1, value=DOC_TITLES[doc.doc_type])
    subtitle_cell.font = Font(bold=True)
    subtitle_cell.alignment = Alignment(horizontal="center")

    # Statement metadata; every doc type has at least four columns here.
    sheet.cell(row=3, column=1, value="Statement No:")
    sheet.cell(row=3, column=2, value=doc.statement_number)
    sheet.cell(row=3, column=3, value="Statement Date:")
    sheet.cell(row=3, column=4, value=statement_date)

    # Two-row header: group row over the field-synonym row.
    group_row, field_row = 5, 6
    reference_span = column_count - 3
    group_spans = (
        (1, reference_span, XLSX_GROUP_LABELS[0]),
        (reference_span + 1, reference_span + 2, XLSX_GROUP_LABELS[1]),
        (column_count, column_count, XLSX_GROUP_LABELS[2]),
    )
    for start, end, label in group_spans:
        if end > start:
            sheet.merge_cells(
                start_row=group_row, start_column=start, end_row=group_row, end_column=end
            )
        group_cell = sheet.cell(row=group_row, column=start, value=label)
        group_cell.font = Font(bold=True)
        group_cell.alignment = Alignment(horizontal="center")
    for index, field in enumerate(fields, start=1):
        sheet.cell(row=field_row, column=index, value=synonyms[field]).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(index)].width = XLSX_COLUMN_WIDTHS[field]

    lines = sorted(doc.lines, key=lambda item: item.line_no)
    totals_after = max(1, len(lines) if totals_at_end else len(lines) // 2)
    row_index = field_row
    amount_column = column_count
    totals_written = False
    for position, line in enumerate(lines, start=1):
        row_index += 1
        for index, field in enumerate(fields, start=1):
            if field is FieldName.LINE_AMOUNT and rng.random() < XLSX_CURRENCY_CELL_RATIO:
                # Currency-formatted numeric cell; record what the format displays.
                cell = sheet.cell(
                    row=row_index, column=index, value=line.amount_cents / CENTS_PER_DOLLAR
                )
                cell.number_format = XLSX_CURRENCY_FORMAT
                value = format_amount(AmountStyle.DOLLAR_COMMA, line.amount_cents)
            else:
                value = _line_cell(line, field, amount_style, date_style)
                sheet.cell(row=row_index, column=index, value=value)
            if value:
                rendered[line_key(line.line_no, field)] = value
        if position == totals_after:
            row_index = _write_totals_row(sheet, row_index, amount_column, doc.subtotal_cents)
            totals_written = True
    if not totals_written:
        _write_totals_row(sheet, row_index, amount_column, doc.subtotal_cents)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(out_path)
    _pin_workbook_timestamps(out_path)
    return rendered


def _pin_workbook_timestamps(path: Path) -> None:
    """Rewrite the saved workbook so its bytes carry no wall-clock time."""
    with zipfile.ZipFile(path) as source:
        members = [(info, source.read(info.filename)) for info in source.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in members:
            data = (
                _CORE_TIMESTAMP_ELEMENT.sub(rb"\g<1>" + _CORE_TIMESTAMP_TEXT + rb"\g<2>", payload)
                if info.filename == XLSX_CORE_PROPERTIES
                else payload
            )
            pinned = zipfile.ZipInfo(info.filename, date_time=XLSX_FIXED_ZIP_TIME)
            pinned.compress_type = info.compress_type
            pinned.create_system = info.create_system
            pinned.external_attr = info.external_attr
            target.writestr(pinned, data)
