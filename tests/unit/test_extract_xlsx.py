"""The spreadsheet extractor: header discovery, exact cents, and its limits.

Every workbook here is built offline, either by the dataset renderer itself or
cell by cell, so the suite never reaches for a fixture file.
"""

import zipfile
from collections.abc import Sequence
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from crossfoot.constants import (
    DocType,
    ExtractionRoute,
    FieldName,
    IngestErrorKind,
    LineType,
    Oem,
    QualityTier,
    ReviewStatus,
)
from crossfoot.extraction import xlsx as xlsx_module
from crossfoot.extraction.tabular import MAX_CELL_CHARS, MAX_DATA_ROWS, MAX_FILE_BYTES
from crossfoot.extraction.xlsx import extract_xlsx
from crossfoot.generator.renderers.tabular import XLSX_CURRENCY_FORMAT, render_xlsx
from crossfoot.models.extraction import ExtractedDocument, ExtractedField
from crossfoot.models.statement import StatementDoc, StatementLine

DOC_ID = "doc-xlsx-01"
TEMPLATE_ID = "northstar-warranty_credit_memo-xlsx-v1"
SEED = 7

# Amounts chosen so the assertions cannot pass by accident: 1234.55 and 7.07 are
# both unrepresentable in binary floating point, so any float arithmetic on the
# way out of openpyxl shows up as a cent of drift.
STATEMENT = StatementDoc(
    doc_id=DOC_ID,
    dealer_id="dlr-northstar",
    doc_type=DocType.WARRANTY_CREDIT_MEMO,
    oem=Oem.NORTHSTAR,
    statement_number="WC-2026-07-114",
    statement_date=date(2026, 7, 31),
    period_start=date(2026, 7, 1),
    period_end=date(2026, 7, 31),
    subtotal_cents=25_397,
    total_cents=25_397,
    lines=(
        StatementLine(
            line_no=1,
            line_type=LineType.CHARGE,
            claim_number="NS10000001",
            ro_number="100001",
            vin="1N4AL3AP7DC900001",
            line_date=date(2026, 7, 10),
            description="Brake caliper reseal",
            amount_cents=123_455,
        ),
        StatementLine(
            line_no=2,
            line_type=LineType.CHARGE,
            claim_number="NS10000002",
            ro_number="100002",
            vin="1N4AL3AP7DC900002",
            line_date=date(2026, 7, 18),
            description="Cabin filter replacement",
            amount_cents=707,
        ),
        StatementLine(
            line_no=3,
            line_type=LineType.CREDIT,
            claim_number="NS10000003",
            ro_number="100003",
            vin="1N4AL3AP7DC900003",
            line_date=date(2026, 7, 25),
            description="Core return credit",
            amount_cents=-98_765,
        ),
    ),
)

SINGLE_HEADER: tuple[object, ...] = ("Claim Number", "Date", "Description", "Amount")
DATA_ROW: tuple[object, ...] = ("NS12345678", "07/15/2026", "Alpha brake kit", "123.45")
GIANT_CELL_MARKER = "giant-cell-marker"


def _write_workbook(
    path: Path,
    rows: Sequence[Sequence[object]],
    *,
    merges: Sequence[str] = (),
    currency_cells: Sequence[str] = (),
) -> Path:
    """A workbook written cell by cell, so one layout hazard can be isolated."""
    book = Workbook()
    sheet = book.active
    for row in rows:
        sheet.append(list(row))
    for span in merges:
        sheet.merge_cells(span)
    for reference in currency_cells:
        sheet[reference].number_format = XLSX_CURRENCY_FORMAT
    book.save(path)
    return path


def _inflate_cell(path: Path, size: int) -> Path:
    """Rewrite the marker cell past Excel's own 32767 character ceiling.

    No writer emits a cell this long, openpyxl included, so a hand-edited
    container is the only way to reach the guard that keeps one absurd shared
    string out of an extracted field.
    """
    with zipfile.ZipFile(path) as source:
        members = [(info.filename, source.read(info.filename)) for info in source.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, payload in members:
            target.writestr(name, payload.replace(GIANT_CELL_MARKER.encode(), b"x" * size))
    return path


def _by_line(doc: ExtractedDocument, line_no: int) -> dict[FieldName, ExtractedField]:
    return {field.name: field for field in doc.line_fields if field.line_no == line_no}


def _cents(doc: ExtractedDocument) -> list[int | None]:
    return [field.value_cents for field in doc.line_fields if field.name is FieldName.LINE_AMOUNT]


# The workbook the dataset generator actually produces.


def test_generated_workbook_extracts_every_line_with_exact_cents(tmp_path: Path) -> None:
    path = tmp_path / "statement.xlsx"
    render_xlsx(STATEMENT, TEMPLATE_ID, SEED, path)
    doc = extract_xlsx(path, DOC_ID)
    assert doc.route is ExtractionRoute.XLSX
    assert doc.error is None
    assert {field.line_no for field in doc.line_fields} == {1, 2, 3}
    for line in STATEMENT.lines:
        fields = _by_line(doc, line.line_no)
        assert fields[FieldName.LINE_AMOUNT].value_cents == line.amount_cents
        assert fields[FieldName.LINE_DATE].value_date == line.line_date
        assert fields[FieldName.DESCRIPTION].value == line.description
        assert fields[FieldName.CLAIM_NUMBER].value == line.claim_number
        assert fields[FieldName.VIN].value == line.vin


def test_generated_workbook_line_numbers_follow_sheet_order(tmp_path: Path) -> None:
    path = tmp_path / "ordinals.xlsx"
    render_xlsx(STATEMENT, TEMPLATE_ID, SEED, path)
    doc = extract_xlsx(path, DOC_ID)
    ordered = [
        (field.line_no, field.value)
        for field in doc.line_fields
        if field.name is FieldName.DESCRIPTION
    ]
    assert ordered == [(line.line_no, line.description) for line in STATEMENT.lines]


def test_generated_workbook_fields_carry_the_spreadsheet_baseline(tmp_path: Path) -> None:
    path = tmp_path / "signals.xlsx"
    render_xlsx(STATEMENT, TEMPLATE_ID, SEED, path)
    doc = extract_xlsx(path, DOC_ID)
    amount = _by_line(doc, 1)[FieldName.LINE_AMOUNT]
    assert amount.signals.quality_tier is QualityTier.XLSX
    assert amount.signals.validator_pass == 1.0
    assert amount.confidence == 1.0
    assert amount.status is ReviewStatus.AUTO_ACCEPTED
    assert _by_line(doc, 1)[FieldName.CLAIM_NUMBER].signals.grammar_match == 1.0


# Header discovery under the layouts a spreadsheet export drops on it.


def test_two_row_header_maps_every_column(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "two_row_header.xlsx",
        [
            ("References", None, "Post", "Activity", "Amounts"),
            ("Claim Number", "RO #", "Dt", "Description", "Amount"),
            ("NS12345678", "123456", "07/15/2026", "Alpha brake kit", 1234.55),
        ],
        merges=["A1:B1"],
        currency_cells=["E3"],
    )
    fields = _by_line(extract_xlsx(path, DOC_ID), 1)
    assert set(fields) == {
        FieldName.CLAIM_NUMBER,
        FieldName.RO_NUMBER,
        FieldName.LINE_DATE,
        FieldName.DESCRIPTION,
        FieldName.LINE_AMOUNT,
    }
    assert fields[FieldName.CLAIM_NUMBER].value == "NS12345678"
    assert fields[FieldName.RO_NUMBER].value == "123456"
    assert fields[FieldName.LINE_DATE].value_date == date(2026, 7, 15)
    assert fields[FieldName.DESCRIPTION].value == "Alpha brake kit"
    assert fields[FieldName.LINE_AMOUNT].value_cents == 123_455


def test_merged_title_block_is_not_mistaken_for_a_header(tmp_path: Path) -> None:
    """A one-column title reading "Amount" is decoration, not a header row."""
    path = _write_workbook(
        tmp_path / "merged_title.xlsx",
        [
            ("Amount", None, None, None),
            ("Warranty Credit Memo", None, None, None),
            SINGLE_HEADER,
            DATA_ROW,
        ],
        merges=["A1:D1", "A2:D2"],
    )
    doc = extract_xlsx(path, DOC_ID)
    assert {field.line_no for field in doc.line_fields} == {1}
    fields = _by_line(doc, 1)
    assert fields[FieldName.CLAIM_NUMBER].value == "NS12345678"
    assert fields[FieldName.LINE_AMOUNT].value_cents == 12_345


def test_totals_row_is_skipped(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "totals.xlsx",
        [
            SINGLE_HEADER,
            ("NS12345678", "07/15/2026", "Alpha brake kit", "10.00"),
            ("TOTALS", None, None, 30.0),
            ("NS87654321", "07/16/2026", "Bravo core return", "20.00"),
        ],
        currency_cells=["D3"],
    )
    doc = extract_xlsx(path, DOC_ID)
    assert {field.line_no for field in doc.line_fields} == {1, 2}
    assert _cents(doc) == [1_000, 2_000]
    assert all(field.raw_text != "TOTALS" for field in doc.line_fields)


def test_blank_rows_take_no_line_number(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "blank_rows.xlsx",
        [
            SINGLE_HEADER,
            ("NS12345678", "07/15/2026", "Alpha brake kit", "10.00"),
            (None, None, None, None),
            ("NS87654321", "07/16/2026", "Bravo core return", "20.00"),
        ],
    )
    doc = extract_xlsx(path, DOC_ID)
    assert {field.line_no for field in doc.line_fields} == {1, 2}


def test_headerless_workbook_is_a_typed_error(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "headerless.xlsx",
        [("nothing", "useful", "here"), ("still", "nothing", "useful")],
    )
    doc = extract_xlsx(path, DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.UNRECOGNIZED
    assert doc.line_fields == ()


# Exact cents out of a numeric cell.


@pytest.mark.parametrize(
    ("stored", "expected_cents", "expected_raw"),
    [
        (1234.55, 123_455, "$1,234.55"),
        (0.07, 7, "$0.07"),
        (-987.65, -98_765, "-$987.65"),
        (1_234_567.89, 123_456_789, "$1,234,567.89"),
    ],
)
def test_currency_cell_recovers_exact_cents(
    tmp_path: Path, stored: float, expected_cents: int, expected_raw: str
) -> None:
    path = _write_workbook(
        tmp_path / "currency.xlsx",
        [SINGLE_HEADER, ("NS12345678", "07/15/2026", "Alpha brake kit", stored)],
        currency_cells=["D2"],
    )
    amount = _by_line(extract_xlsx(path, DOC_ID), 1)[FieldName.LINE_AMOUNT]
    assert amount.value_cents == expected_cents
    assert amount.raw_text == expected_raw


def test_unformatted_numeric_cell_recovers_exact_cents(tmp_path: Path) -> None:
    """No currency format, so the cell reads as the plain digits it stores."""
    path = _write_workbook(
        tmp_path / "plain_numeric.xlsx",
        [SINGLE_HEADER, ("NS12345678", "07/15/2026", "Alpha brake kit", 1234.55)],
    )
    amount = _by_line(extract_xlsx(path, DOC_ID), 1)[FieldName.LINE_AMOUNT]
    assert amount.raw_text == "1234.55"
    assert amount.value_cents == 123_455


# Resource ceilings, shared with the CSV path.


def test_oversize_file_is_refused_by_stat(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    path = tmp_path / "huge.xlsx"
    with path.open("wb") as handle:
        handle.truncate(MAX_FILE_BYTES + 1)

    def _forbidden(*args: object, **kwargs: object) -> object:
        raise AssertionError("oversize workbook must be refused before it is opened")

    monkeypatch.setattr(xlsx_module, "load_workbook", _forbidden)
    doc = extract_xlsx(path, DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.TOO_LARGE


def test_row_cap_is_enforced(tmp_path: Path) -> None:
    rows = [SINGLE_HEADER, *[DATA_ROW] * (MAX_DATA_ROWS + 1)]
    doc = extract_xlsx(_write_workbook(tmp_path / "too_many_rows.xlsx", rows), DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.TOO_LARGE
    assert str(MAX_DATA_ROWS) in doc.error.detail


def test_giant_cell_is_a_typed_error_not_an_exception(tmp_path: Path) -> None:
    row = ("NS12345678", "07/15/2026", GIANT_CELL_MARKER, "123.45")
    path = _write_workbook(tmp_path / "giant_cell.xlsx", [SINGLE_HEADER, row])
    doc = extract_xlsx(_inflate_cell(path, MAX_CELL_CHARS + 1), DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.TOO_LARGE


# Input that is not a workbook at all.


def test_delimited_text_under_a_workbook_name_is_a_typed_error(tmp_path: Path) -> None:
    path = tmp_path / "actually_csv.xlsx"
    path.write_bytes(b"Claim Number,Date,Description,Amount\nNS12345678,07/15/2026,Alpha,1.00\n")
    doc = extract_xlsx(path, DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.UNRECOGNIZED


def test_truncated_workbook_is_a_typed_error(tmp_path: Path) -> None:
    whole = _write_workbook(tmp_path / "whole.xlsx", [SINGLE_HEADER, DATA_ROW])
    path = tmp_path / "truncated.xlsx"
    path.write_bytes(whole.read_bytes()[:512])
    doc = extract_xlsx(path, DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.UNRECOGNIZED


def test_plain_zip_is_a_typed_error(tmp_path: Path) -> None:
    path = tmp_path / "archive.xlsx"
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("notes.txt", "not a workbook")
    doc = extract_xlsx(path, DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.UNRECOGNIZED


def test_missing_file_is_a_typed_error(tmp_path: Path) -> None:
    doc = extract_xlsx(tmp_path / "absent.xlsx", DOC_ID)
    assert doc.route is ExtractionRoute.UNPROCESSABLE
    assert doc.error is not None
    assert doc.error.kind is IngestErrorKind.UNRECOGNIZED
