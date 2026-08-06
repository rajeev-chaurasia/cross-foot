"""Rendered-values completeness and template health for every renderer."""

import time
import zipfile
from collections.abc import Iterator
from datetime import date
from pathlib import Path

import pdfplumber
import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]
from pypdf import PdfReader

from crossfoot.constants import CSV_HEADER_SYNONYMS, DocType, FieldName, LineType, Oem
from crossfoot.generator.renderers.base import (
    DOC_LINE_FIELDS,
    ISSUER_ADDRESSES,
    MARQUE_BRANDING,
    PAYABLE_DOC_TYPES,
    REMIT_ADDRESSES,
    format_due_date,
    header_key,
    line_key,
)
from crossfoot.generator.renderers.chromium import (
    PDF_FIXED_DATE,
    ChromiumPdfRenderer,
    build_context,
    render_html,
)
from crossfoot.generator.renderers.tabular import (
    XLSX_CORE_PROPERTIES,
    XLSX_FIXED_ZIP_TIME,
    render_csv,
    render_xlsx,
)
from crossfoot.models.statement import StatementDoc, StatementLine

RENDER_SEED = 11
# Long enough for a wall-clock stamp to change between two renders.
CLOCK_TICK_SECONDS = 1.5
# A one-page statement prints far more than this; a blank page prints less.
MIN_PDF_CHARS = 200

# The composer gives these doc types a previous balance; the others carry none.
BALANCE_FORWARD_DOC_TYPES = frozenset({DocType.PARTS_STATEMENT, DocType.FLOORPLAN_STATEMENT})

# Known-good ISO 3779 VIN (check digit X at position 9).
SAMPLE_VIN = "1M8GDM9AXKP042788"


def make_doc(oem: Oem, doc_type: DocType) -> StatementDoc:
    """Two-line statement respecting the crossfoot invariant, all references set."""
    lines = (
        StatementLine(
            line_no=1,
            line_type=LineType.CHARGE,
            claim_number="1234A56789",
            ro_number="RO123456",
            vin=SAMPLE_VIN,
            invoice_number="M1234567",
            program_code="PGM-0001",
            line_date=date(2026, 7, 6),
            description="Alpha brake kit restock",
            amount_cents=123_456,
            source_entry_id="led-parts_payable-00001",
        ),
        StatementLine(
            line_no=2,
            line_type=LineType.CREDIT,
            claim_number="1234A56790",
            ro_number="RO123457",
            vin=SAMPLE_VIN,
            invoice_number="M1234568",
            program_code="PGM-0002",
            line_date=date(2026, 7, 21),
            description="Bravo core return credit",
            amount_cents=-2_050,
            source_entry_id="led-parts_payable-00002",
        ),
    )
    subtotal_cents = sum(line.amount_cents for line in lines)
    # Mirror the composer: only balance-forward doc types carry a previous balance.
    previous_balance_cents = 55_000 if doc_type in BALANCE_FORWARD_DOC_TYPES else None
    return StatementDoc(
        doc_id=f"doc-{doc_type}-dlr-{oem}-202607-01",
        dealer_id=f"dlr-{oem}",
        doc_type=doc_type,
        oem=oem,
        statement_number="STMT-202607-01",
        statement_date=date(2026, 7, 31),
        period_start=date(2026, 7, 1),
        period_end=date(2026, 7, 31),
        previous_balance_cents=previous_balance_cents,
        subtotal_cents=subtotal_cents,
        adjustments_cents=0,
        total_cents=(previous_balance_cents or 0) + subtotal_cents,
        lines=lines,
    )


def _expected_line_keys(doc: StatementDoc) -> set[str]:
    return {
        line_key(line.line_no, field)
        for line in doc.lines
        for field in DOC_LINE_FIELDS[doc.doc_type]
    }


def _expected_pdf_keys(doc: StatementDoc) -> set[str]:
    keys = _expected_line_keys(doc) | {
        header_key(FieldName.STATEMENT_NUMBER),
        header_key(FieldName.STATEMENT_DATE),
        header_key(FieldName.SUBTOTAL),
        header_key(FieldName.TOTAL),
    }
    if doc.previous_balance_cents is not None:
        keys.add(header_key(FieldName.PREVIOUS_BALANCE))
    return keys


@pytest.fixture(scope="module")
def pdf_renderer() -> Iterator[ChromiumPdfRenderer]:
    """One browser for the whole template matrix."""
    with ChromiumPdfRenderer() as renderer:
        yield renderer


def _pdf_chars_and_text(path: Path) -> tuple[int, str]:
    with pdfplumber.open(path) as pdf:
        chars = sum(len(page.chars) for page in pdf.pages)
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    return chars, text


def _squashed(text: str) -> str:
    """Case-folded and whitespace-free, so CSS uppercasing, letter-spacing, and
    column wrapping do not turn a printed value into a false negative."""
    return "".join(text.split()).casefold()


@pytest.mark.parametrize("oem", tuple(Oem))
@pytest.mark.parametrize("doc_type", tuple(DocType))
def test_every_template_renders_all_values(oem: Oem, doc_type: DocType) -> None:
    doc = make_doc(oem, doc_type)
    html, rendered = render_html(doc, f"{oem}-{doc_type}-v1")
    assert set(rendered) == _expected_pdf_keys(doc)
    for key, value in rendered.items():
        assert value in html, f"{key} value {value!r} missing from HTML"
    assert MARQUE_BRANDING[oem].name in html


@pytest.mark.parametrize("oem", tuple(Oem))
@pytest.mark.parametrize("doc_type", tuple(DocType))
def test_every_template_prints_a_readable_pdf(
    oem: Oem, doc_type: DocType, tmp_path: Path, pdf_renderer: ChromiumPdfRenderer
) -> None:
    """All 16 (marque, doc_type) templates print text-bearing PDFs."""
    doc = make_doc(oem, doc_type)
    out_path = tmp_path / f"{oem}-{doc_type}.pdf"
    rendered = pdf_renderer.render(doc, f"{oem}-{doc_type}-v1", RENDER_SEED, out_path)

    assert set(rendered) == _expected_pdf_keys(doc)
    chars, text = _pdf_chars_and_text(out_path)
    assert chars > MIN_PDF_CHARS, f"{oem} {doc_type}: only {chars} extractable chars"
    printed = _squashed(text)
    for key, value in rendered.items():
        assert _squashed(value) in printed, f"{oem} {doc_type}: {key} value {value!r} not printed"
    assert _squashed(MARQUE_BRANDING[oem].name) in printed


@pytest.mark.parametrize("oem", tuple(Oem))
@pytest.mark.parametrize("doc_type", tuple(PAYABLE_DOC_TYPES))
def test_payable_doc_types_carry_due_date_and_lockbox(oem: Oem, doc_type: DocType) -> None:
    doc = make_doc(oem, doc_type)
    context, _ = build_context(doc)
    assert context["due_date"] == format_due_date(oem, doc.statement_date)
    assert context["remit_address"] == REMIT_ADDRESSES[oem, doc_type]


@pytest.mark.parametrize("doc_type", (DocType.WARRANTY_CREDIT_MEMO, DocType.INCENTIVE_STATEMENT))
def test_receivable_doc_types_omit_remittance_values(doc_type: DocType) -> None:
    context, _ = build_context(make_doc(Oem.ATLAS, doc_type))
    assert "due_date" not in context
    assert "remit_address" not in context


@pytest.mark.parametrize("doc_type", tuple(DocType))
def test_atlas_prints_its_per_doc_type_issuer_address(doc_type: DocType) -> None:
    context, _ = build_context(make_doc(Oem.ATLAS, doc_type))
    assert context["marque_address"] == ISSUER_ADDRESSES[Oem.ATLAS, doc_type]


def test_render_csv_covers_lines_in_truth_order(tmp_path: Path) -> None:
    doc = make_doc(Oem.MERIDIAN, DocType.PARTS_STATEMENT)
    out_path = tmp_path / "parts.csv"
    rendered = render_csv(doc, "meridian-parts_statement-csv-v1", RENDER_SEED, out_path)
    assert set(rendered) == _expected_line_keys(doc)
    # All content is ASCII, so both permitted encodings decode identically.
    text = out_path.read_bytes().decode("utf-8")
    first_line = text.splitlines()[0]
    assert any(synonym in first_line for synonym in CSV_HEADER_SYNONYMS[FieldName.INVOICE_NUMBER])
    for key, value in rendered.items():
        assert value in text, f"{key} value {value!r} missing from CSV"
    assert text.index("Alpha brake kit restock") < text.index("Bravo core return credit")


def test_render_csv_is_deterministic_per_seed(tmp_path: Path) -> None:
    doc = make_doc(Oem.NORTHSTAR, DocType.WARRANTY_CREDIT_MEMO)
    first = tmp_path / "first.csv"
    second = tmp_path / "second.csv"
    render_csv(doc, "northstar-warranty_credit_memo-csv-v1", RENDER_SEED, first)
    render_csv(doc, "northstar-warranty_credit_memo-csv-v1", RENDER_SEED, second)
    assert first.read_bytes() == second.read_bytes()


def test_render_xlsx_covers_values_and_layout(tmp_path: Path) -> None:
    doc = make_doc(Oem.NORTHSTAR, DocType.INCENTIVE_STATEMENT)
    out_path = tmp_path / "incentive.xlsx"
    rendered = render_xlsx(doc, "northstar-incentive_statement-xlsx-v1", RENDER_SEED, out_path)
    expected = _expected_line_keys(doc) | {
        header_key(FieldName.STATEMENT_NUMBER),
        header_key(FieldName.STATEMENT_DATE),
        header_key(FieldName.SUBTOTAL),
    }
    assert set(rendered) == expected

    book = load_workbook(out_path)
    sheet = book.active
    assert sheet["A1"].value == MARQUE_BRANDING[Oem.NORTHSTAR].name
    assert sheet["B3"].value == doc.statement_number
    # XLSX orders reference columns first; incentive leads with the program code.
    assert sheet["A6"].value in CSV_HEADER_SYNONYMS[FieldName.PROGRAM_CODE]
    first_column = [row[0].value for row in sheet.iter_rows(min_col=1, max_col=1)]
    assert "TOTALS" in first_column
    all_values = {cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None}
    assert "Alpha brake kit restock" in all_values


def test_render_xlsx_is_byte_identical_across_the_clock(tmp_path: Path) -> None:
    doc = make_doc(Oem.KAIZEN, DocType.PARTS_STATEMENT)
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    render_xlsx(doc, "kaizen-parts_statement-xlsx-v1", RENDER_SEED, first)
    time.sleep(CLOCK_TICK_SECONDS)
    render_xlsx(doc, "kaizen-parts_statement-xlsx-v1", RENDER_SEED, second)
    assert first.read_bytes() == second.read_bytes()


def test_render_xlsx_pins_zip_and_core_timestamps(tmp_path: Path) -> None:
    doc = make_doc(Oem.KAIZEN, DocType.PARTS_STATEMENT)
    out_path = tmp_path / "pinned.xlsx"
    render_xlsx(doc, "kaizen-parts_statement-xlsx-v1", RENDER_SEED, out_path)
    with zipfile.ZipFile(out_path) as archive:
        assert {info.date_time for info in archive.infolist()} == {XLSX_FIXED_ZIP_TIME}
        core = archive.read(XLSX_CORE_PROPERTIES).decode("utf-8")
    assert core.count("2026-01-01T00:00:00Z") == 2  # dcterms:created and dcterms:modified


def test_chromium_pdf_is_byte_identical_across_the_clock(
    tmp_path: Path, pdf_renderer: ChromiumPdfRenderer
) -> None:
    doc = make_doc(Oem.MERIDIAN, DocType.PARTS_STATEMENT)
    first = tmp_path / "first.pdf"
    second = tmp_path / "second.pdf"
    pdf_renderer.render(doc, "meridian-parts_statement-v1", RENDER_SEED, first)
    time.sleep(CLOCK_TICK_SECONDS)
    pdf_renderer.render(doc, "meridian-parts_statement-v1", RENDER_SEED, second)
    assert first.read_bytes() == second.read_bytes()
    metadata = PdfReader(first).metadata
    assert metadata is not None
    assert str(metadata["/CreationDate"]) == PDF_FIXED_DATE
    assert str(metadata["/ModDate"]) == PDF_FIXED_DATE
